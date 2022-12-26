import openpyxl

file="C:\\Users\\DHARUN R\\Desktop\\sample.xlsx"

workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet1"]

rows=sheet.max_row
column=sheet.max_column

for r in range(1,rows+1):
    for c in range(1,column+1):
        print(sheet.cell(r,c).value,end='       ')
    print("\n")






